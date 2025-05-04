import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from PyPDF2 import PdfMerger, PdfReader

st.set_page_config(page_title="Brunch Formatter", layout="centered")
st.title("🥂 Brunch Formatter")

uploaded_file = st.file_uploader("Upload booking CSV", type="csv")
double_sided = st.checkbox("Generate double-sided reservation cards", value=False)
generate_button = st.button("Generate Excel + PDF")

def extract_table_numbers(area_value):
    if isinstance(area_value, str):
        matches = re.findall(r"(?:Wilsons?|Wilson's)\s*(\d+[a-zA-Z]?)", area_value, re.IGNORECASE)
        cleaned = ["STAGE" if m == "3" else m for m in matches]
        return ", ".join(cleaned) if cleaned else "TBC"
    return "TBC"

def extract_deposit(value):
    if isinstance(value, str):
        match = re.search(r"£\s?(\d+(?:\.\d{1,2})?)", value)
        return float(match.group(1)) if match else 0.0
    return float(value) if isinstance(value, (int, float)) else 0.0

def create_excel(df):
    wb = Workbook()
    ws = wb.active

    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    red_font = Font(color="FF0000")
    header_font = Font(bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = center_align
            if r_idx == 1:
                cell.font = header_font
            elif df.columns[c_idx - 1] == "AMOUNT DUE:":
                if isinstance(value, str) and value.startswith("£"):
                    try:
                        if float(value[1:]) > 0:
                            cell.font = red_font
                    except:
                        pass

    col_widths = {
        "GUESTS": 7,
        "TIME": 7,
        "PRE-PAYMENT:": 10,
        "AMOUNT DUE:": 10,
        "LAST ORDERS:": 10,
        "FREE SHOTS?": 6,
        "TIME TABLE IS NEEDED BACK:": 15,
        "FLIP TIME": 8,
        "CLEAR ORDER": 10
    }

    for idx, col in enumerate(df.columns, 1):
        letter = ws.cell(row=1, column=idx).column_letter
        width = col_widths.get(col.upper())
        if width:
            ws.column_dimensions[letter].width = width
        else:
            ws.column_dimensions[letter].width = 40 if col.upper() == "RUN SHEET NOTES:" else 20

    return wb

def create_front(name, time_range):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    name_x, name_y = 105 * mm, 95 * mm
    time_x, time_y = 105 * mm, 30 * mm
    max_width_pt = 157.5 * mm

    name_font = "Courier-BoldOblique"
    name_size = 60
    while c.stringWidth(name, name_font, name_size) > max_width_pt:
        name_size -= 1
    c.setFont(name_font, name_size)
    c.drawCentredString(name_x, name_y, name)

    time_font = "Courier-BoldOblique"
    time_size = 60
    while c.stringWidth(time_range, time_font, time_size) > max_width_pt and time_size > 40:
        time_size -= 1
    c.setFont(time_font, time_size)
    c.drawCentredString(time_x, time_y, time_range)

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer

def create_back(table, guests):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    x, y = 10 * mm, 297 * mm - 10 * mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(x, y, f"Table: {table}")
    c.drawString(x, y - 18, f"Guests: {guests}")
    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer

def generate_outputs(upload):
    lines = upload.getvalue().decode("utf-8").splitlines()
    header_row_index = next(i for i, line in enumerate(lines) if "Time" in line and "Guests" in line)
    df = pd.read_csv(BytesIO(upload.getvalue()), skiprows=header_row_index)
    df.columns = [col.strip() for col in df.columns]

    df["TABLE"] = df["Area"].apply(extract_table_numbers)
    prepayment_col = next((col for col in df.columns if any(k in col.lower() for k in ["pre", "paid", "deposit"])), df.columns[-1])
    df["Cleaned Deposits"] = df["Deposits"].apply(extract_deposit)

    formatted = pd.DataFrame()
    formatted["NAME"] = df[df.columns[0]]
    formatted["GUESTS"] = df["Guests"]
    formatted["TIME"] = df["Time"]
    formatted["TABLE"] = df["TABLE"]
    formatted["PRE-PAYMENT:"] = df["Cleaned Deposits"].apply(lambda x: f"£{x:.2f}")
    formatted["AMOUNT DUE:"] = (
        pd.to_numeric(df["Guests"], errors="coerce").fillna(0) * 39.5 - df["Cleaned Deposits"]
    ).apply(lambda x: "-" if x <= 0 else f"£{x:.2f}")
    formatted["LAST ORDERS:"] = pd.to_datetime(df["Time"], format="%H:%M", errors="coerce").apply(
        lambda t: (t + timedelta(minutes=75)).strftime("%H:%M") if pd.notnull(t) else ""
    )
    
    df["RUN SHEET NOTES:"] = df[["Customer preorder notes", "Customer requests", "Dietary requirements"]].apply(lambda row: "\n".join([str(val).strip() for val in row if pd.notna(val) and str(val).strip() != ""]), axis=1)
    formatted["RUN SHEET NOTES:"] = df["RUN SHEET NOTES:"]
    formatted["TIME TABLE IS NEEDED BACK:"] = ""
    formatted["FLIP TIME"] = ""
    formatted["CLEAR ORDER"] = ""
    formatted["FREE SHOTS?"] = ""

    formatted["START"] = pd.to_datetime(formatted["TIME"], format="%H:%M", errors="coerce")
    formatted["END"] = formatted["START"] + timedelta(minutes=90)

    time_back_list = []
    flip_time_list = []

    for idx, row in formatted.iterrows():
        table = row["TABLE"]
        end_time = row["END"]
        others = formatted[(formatted["TABLE"] == table) & (formatted.index != idx)]
        future = others[others["START"] > row["START"]]

        if not future.empty:
            next_start = future["START"].min()
            gap = (next_start - end_time).total_seconds() / 60
            if gap <= 15:
                needed_back = end_time
            elif gap <= 30:
                needed_back = end_time + timedelta(minutes=10)
            else:
                needed_back = next_start - timedelta(minutes=30)
            flip = (next_start - needed_back).total_seconds() / 60
            flip_time_str = f"{int(flip)} mins"
        else:
            needed_back = end_time
            flip_time_str = ""

        time_back_list.append(needed_back.strftime("%H:%M") if pd.notnull(needed_back) else "")
        flip_time_list.append(flip_time_str)

    formatted["TIME TABLE IS NEEDED BACK:"] = time_back_list
    formatted["FLIP TIME"] = flip_time_list

    flip_dt = pd.to_datetime(formatted["TIME TABLE IS NEEDED BACK:"], format="%H:%M", errors="coerce")
    order_map = {t: i + 1 for i, t in enumerate(sorted(flip_dt.dropna().unique()))}
    formatted["CLEAR ORDER"] = flip_dt.map(order_map).fillna("").astype(str)

    formatted["SORT_TIME"] = pd.to_datetime(formatted["TIME"], format="%H:%M", errors="coerce")
    formatted = formatted.sort_values(by="SORT_TIME").drop(columns=["START", "END", "SORT_TIME"])

    wb = create_excel(formatted)
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    merger = PdfMerger()
    for _, row in formatted.iterrows():
        name = str(row["NAME"]).strip()
        guests = str(row["GUESTS"]).strip()
        table = str(row["TABLE"]).strip()
        start_str = str(row["TIME"]).strip()
        try:
            start = datetime.strptime(start_str, "%H:%M")
            end = start + timedelta(minutes=90)
            time_range = f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}"
        except:
            time_range = start_str

        front = PdfReader(create_front(name, time_range), strict=False)
        merger.append(front)
        if double_sided:
            back = PdfReader(create_back(table, guests), strict=False)
            merger.append(back)

    pdf_bytes = BytesIO()
    merger.write(pdf_bytes)
    merger.close()
    pdf_bytes.seek(0)

    return excel_bytes, pdf_bytes

if generate_button and uploaded_file:
    with st.spinner("Generating files..."):
        excel_file, pdf_file = generate_outputs(uploaded_file)
        st.session_state["excel_file"] = excel_file
        st.session_state["pdf_file"] = pdf_file
    st.success("✅ Files ready!")

if "excel_file" in st.session_state:
    st.download_button(
        "📥 Download Excel", 
        data=st.session_state["excel_file"], 
        file_name="brunch_sheet.xlsx",
        key="download_excel"
    )

if "pdf_file" in st.session_state:
    st.download_button(
        "📥 Download PDF", 
        data=st.session_state["pdf_file"], 
        file_name="reservation_cards.pdf",
        key="download_pdf"
    )
