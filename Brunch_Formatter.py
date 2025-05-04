import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import re
import os
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from PyPDF2 import PdfMerger, PdfReader
from io import BytesIO

# ---------- Settings Path ----------
SETTINGS_FILE = os.path.join(os.path.expanduser("~"), ".brunch_formatter_paths.json")

def load_last_paths():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, "r") as f:
            return json.load(f)
    return {}

def save_last_paths(paths):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(paths, f)

# ---------- Helpers ----------
def extract_table_numbers(area_value):
    if isinstance(area_value, str):
        matches = re.findall(r"(?:Wilsons?|Wilson's)\s*(\d+[a-zA-Z]?)", area_value, re.IGNORECASE)
        cleaned = ["STAGE" if m == "3" else m for m in matches]
        return ", ".join(cleaned) if cleaned else "TBC"
    return "TBC"

def extract_deposit(value):
    if isinstance(value, str):
        match = re.search(r"£\s?(\d+(?:\.\d{1,2})?)", value)
        return float(match.group(1)) if match else 0.0
    return float(value) if isinstance(value, (int, float)) else 0.0

# ---------- Excel Formatter ----------
def format_brunch_sheet(input_csv, output_excel):
    with open(input_csv, encoding='utf-8') as f:
        lines = f.readlines()
    header_row_index = next(i for i, line in enumerate(lines) if "Time" in line and "Guests" in line)
    df = pd.read_csv(input_csv, skiprows=header_row_index)
    df.columns = [col.strip() for col in df.columns]

    df["TABLE"] = df["Area"].apply(extract_table_numbers)
    df["Cleaned Deposits"] = df[df.columns[-1]].apply(extract_deposit)

    formatted = pd.DataFrame()
    formatted["NAME"] = df[df.columns[0]]
    formatted["GUESTS"] = df["Guests"]
    formatted["TIME"] = df["Time"]
    formatted["TABLE"] = df["TABLE"]
    formatted["Pre-payment:"] = df["Cleaned Deposits"].apply(lambda x: f"£{x:.2f}")
    formatted["Amount Due:"] = (
        pd.to_numeric(df["Guests"], errors="coerce").fillna(0) * 39.5 - df["Cleaned Deposits"]
    ).apply(lambda x: "-" if x <= 0 else f"£{x:.2f}")
    formatted["Last Orders:"] = pd.to_datetime(df["Time"], format="%H:%M", errors="coerce").apply(
        lambda t: (t + timedelta(minutes=75)).strftime("%H:%M") if pd.notnull(t) else ""
    )
    formatted["Run Sheet Notes:"] = df.get("Run Sheet Notes", "")
    formatted["Flip Time"] = ""
    formatted["Clear Order"] = ""
    formatted["FREE SHOTS?"] = ""

    formatted.to_excel(output_excel, index=False)
    wb = load_workbook(output_excel)
    ws = wb.active

    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    red_font = Font(color="FF0000")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = center
            if "Amount Due" in str(ws.cell(row=1, column=cell.column).value):
                if isinstance(cell.value, str) and cell.value.startswith("£"):
                    try:
                        if float(cell.value[1:]) > 0:
                            cell.font = red_font
                    except:
                        pass

    ws.row_dimensions[1].height = 30
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, min(max_len + 2, 30))

    wb.save(output_excel)

# ---------- Reservation Card - Front ----------
def create_front(name, time_range):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    name_x, name_y = 105 * mm, 95 * mm
    time_x, time_y = 105 * mm, 30 * mm
    max_width_pt = 157.5 * mm

    name_font = "Courier-BoldOblique"
    name_size = 60
    while c.stringWidth(name, name_font, name_size) > max_width_pt:
        name_size -= 1
    c.setFont(name_font, name_size)
    c.drawCentredString(name_x, name_y, name)

    time_font = "Courier-BoldOblique"
    time_size = 60
    min_size = 40
    while c.stringWidth(time_range, time_font, time_size) > max_width_pt and time_size > min_size:
        time_size -= 1
    c.setFont(time_font, time_size)
    c.drawCentredString(time_x, time_y, time_range)

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer

# ---------- Reservation Card - Back ----------
def create_back(table, guests):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    x, y = 10 * mm, 297 * mm - 10 * mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(x, y, f"Table: {table}")
    c.drawString(x, y - 18, f"Guests: {guests}")
    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer

# ---------- PDF Generator ----------
def generate_combined_reservation_cards(input_csv, output_pdf, double_sided=True):
    with open(input_csv, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    header_row_index = next(i for i, line in enumerate(lines) if "Time" in line and "Guests" in line)

    df = pd.read_csv(input_csv, skiprows=header_row_index)
    df.columns = [col.strip().upper() for col in df.columns]

    name_col = df.columns[0]
    time_col = "TIME"
    guests_col = "GUESTS"
    table_col = "TABLE" if "TABLE" in df.columns else "AREA"

    if "TABLE" not in df.columns and "AREA" in df.columns:
        df["TABLE"] = df["AREA"].apply(extract_table_numbers)
        table_col = "TABLE"

    merger = PdfMerger()

    for _, row in df.iterrows():
        name = str(row[name_col]).strip()
        start_str = str(row[time_col]).strip()
        try:
            start = datetime.strptime(start_str, "%H:%M")
            end = start + timedelta(minutes=90)
            time_range = f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}"
        except:
            time_range = start_str

        table = str(row[table_col]).strip()
        guests = str(row[guests_col]).strip()

        merger.append(PdfReader(create_front(name, time_range), strict=False))
        if double_sided:
            merger.append(PdfReader(create_back(table, guests), strict=False))

    with open(output_pdf, "wb") as f_out:
        merger.write(f_out)
    merger.close()

# ---------- GUI ----------
def run_gui():
    root = tk.Tk()
    root.title("Brunch Formatter")
    root.geometry("600x440")

    last_paths = load_last_paths()
    input_var = tk.StringVar(value=last_paths.get("input", ""))
    excel_var = tk.StringVar(value=last_paths.get("excel", ""))
    pdf_var = tk.StringVar(value=last_paths.get("pdf", ""))
    double_sided_var = tk.BooleanVar(value=last_paths.get("double_sided", False))

    def browse_input():
        path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")], initialdir=os.path.dirname(input_var.get()) if input_var.get() else None)
        if path:
            input_var.set(path)

    def browse_excel():
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialdir=os.path.dirname(excel_var.get()) if excel_var.get() else None)
        if path:
            excel_var.set(path)

    def browse_pdf():
        path = filedialog.asksaveasfilename(defaultextension=".pdf", initialdir=os.path.dirname(pdf_var.get()) if pdf_var.get() else None)
        if path:
            pdf_var.set(path)

    def generate_both():
        if not input_var.get() or not excel_var.get() or not pdf_var.get():
            messagebox.showerror("Missing", "Please select all file paths.")
            return
        try:
            save_last_paths({
                "input": input_var.get(),
                "excel": excel_var.get(),
                "pdf": pdf_var.get(),
                "double_sided": double_sided_var.get()
            })
            format_brunch_sheet(input_var.get(), excel_var.get())
            generate_combined_reservation_cards(input_var.get(), pdf_var.get(), double_sided=double_sided_var.get())
            messagebox.showinfo("Done", "Excel and PDF created.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # Layout
    tk.Label(root, text="Input CSV:").pack()
    tk.Entry(root, textvariable=input_var, width=60).pack()
    tk.Button(root, text="Browse", command=browse_input).pack(pady=5)

    tk.Label(root, text="Save Excel As:").pack()
    tk.Entry(root, textvariable=excel_var, width=60).pack()
    tk.Button(root, text="Choose Save Location", command=browse_excel).pack(pady=5)

    tk.Label(root, text="Save Reservation PDF:").pack()
    tk.Entry(root, textvariable=pdf_var, width=60).pack()
    tk.Button(root, text="Choose Save Location", command=browse_pdf).pack(pady=5)

    tk.Checkbutton(root, text="Double-sided (front and back)", variable=double_sided_var).pack(pady=5)
    tk.Button(root, text="Generate Both", command=generate_both, width=30).pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    run_gui()
